import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Frutas')

frutas_page = book['Frutas']
frutas_page.append(['Frutas','Quantidade','Preço'])
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Maça','2','R$2,90'])
frutas_page.append(['Pera','3','R$8,90'])

book.save('Planilha de Compras.xlsx')